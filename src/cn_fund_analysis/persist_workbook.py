"""按运行日目录追加写入 Excel：同一天多次运行各占一个 Sheet（Run_001, Run_002, …）。"""

from __future__ import annotations

import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

DEFAULT_WORKBOOK = "fund_metrics_runs.xlsx"

HEADERS = [
    "板块",
    "代码",
    "简称",
    "ret_1m",
    "ret_3m",
    "ret_6m",
    "ret_1y",
    "ret_ytd",
    "年化波动",
    "最大回撤",
]

RATE_COLS = {"ret_1m", "ret_3m", "ret_6m", "ret_1y", "ret_ytd", "年化波动", "最大回撤"}


def _sort_fund_metric_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按板块分组写入：板块顺序与输入行中「首次出现」顺序一致（契合 mapping/JSON 编排）；
    同板块内按代码升序。未见过的新板块排在已出现板块之后。
    """
    order: list[str] = []
    for r in rows:
        s = str(r.get("板块") or "").strip()
        if s not in order:
            order.append(s)
    rank = {s: i for i, s in enumerate(order)}
    return sorted(
        rows,
        key=lambda r: (
            rank.get(str(r.get("板块") or "").strip(), len(order)),
            str(r.get("代码") or "").strip(),
        ),
    )


def ensure_run_date_dir(project_root: Path, run_date: str) -> Path:
    """创建 data/YYYY-MM-DD/（必须存在后才写入工作簿）。"""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", run_date):
        raise ValueError(f"run_date 须为 YYYY-MM-DD，收到: {run_date!r}")
    d = project_root / "data" / run_date
    d.mkdir(parents=True, exist_ok=True)
    return d


def _pct_to_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "")
    try:
        return float(s) / 100.0
    except ValueError:
        return None


def _next_run_sheet_name(wb) -> str:
    nums: list[int] = []
    for name in wb.sheetnames:
        m = re.fullmatch(r"Run_(\d+)", name)
        if m:
            nums.append(int(m.group(1)))
    n = max(nums, default=0) + 1
    return f"Run_{n:03d}"


def append_fund_metrics_sheet(
    project_root: Path,
    run_date: str,
    rows: list[dict[str, Any]],
    *,
    fund_nav_as_of: str,
    topic: str = "",
    workbook_filename: str = DEFAULT_WORKBOOK,
    saved_at: datetime | None = None,
) -> tuple[Path, str]:
    """
    在 data/<run_date>/<workbook_filename> 中追加一个新 Sheet。
    返回 (xlsx 路径, 新 sheet 名)。
    """
    rows = _sort_fund_metric_rows(list(rows))
    day_dir = ensure_run_date_dir(project_root, run_date)
    path = day_dir / workbook_filename

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = _next_run_sheet_name(wb)
    ws = wb.create_sheet(title=sheet_name)

    saved_at = saved_at or datetime.now()
    meta = [
        ("fund_nav_as_of", fund_nav_as_of),
        ("run_folder_date", run_date),
        ("saved_at_local", saved_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("topic", topic or ""),
        ("row_count", str(len(rows))),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    meta_font = Font(bold=True)
    ws.cell(row=1, column=1).font = meta_font

    header_row = len(meta) + 2
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True)

    for r_i, row in enumerate(rows, start=header_row + 1):
        for c_i, h in enumerate(HEADERS, start=1):
            raw = row.get(h)
            cell = ws.cell(row=r_i, column=c_i)
            if h in RATE_COLS:
                fval = _pct_to_float(raw)
                if fval is not None:
                    cell.value = fval
                    cell.number_format = "0.00%"
                else:
                    cell.value = raw
            else:
                cell.value = raw

    wb.save(path)
    return path, sheet_name


def load_rows_from_json(path: Path) -> tuple[list[dict[str, Any]], str, str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = data.get("rows") or data.get("funds")
    if not isinstance(rows, list) or not rows:
        raise ValueError("JSON 须包含非空数组 rows 或 funds")
    as_of = str(data.get("fund_nav_as_of") or data.get("as_of") or "")
    if not as_of:
        raise ValueError("JSON 须包含 fund_nav_as_of 或 as_of")
    topic = str(data.get("topic") or "")
    return rows, as_of, topic


def load_auto_mapping(path: Path) -> tuple[list[dict[str, str]], str]:
    """persist-auto：仅 code + 板块 + 简称；简称可空。"""
    data = json.loads(path.read_text(encoding="utf-8"))
    funds = data.get("funds")
    if not isinstance(funds, list) or not funds:
        raise ValueError("mapping JSON 须包含非空 funds[]")
    topic = str(data.get("topic") or "")
    out: list[dict[str, str]] = []
    for item in funds:
        if not isinstance(item, dict):
            raise ValueError("funds[] 每项须为对象")
        code = str(item.get("code") or item.get("代码") or "").strip()
        if not code:
            raise ValueError("每条 fund 须含 code / 代码")
        sector = str(item.get("板块") or item.get("sector") or "")
        name = str(item.get("简称") or item.get("name") or "")
        out.append({"code": code, "板块": sector, "简称": name})
    return out, topic


def row_from_compute_metrics(
    m: dict[str, Any],
    *,
    sector: str,
    code: str,
    name: str,
) -> dict[str, Any]:
    """将 compute_metrics 字典映射为 persist 宽表一行（值为小数比例，供 Excel 百分比格式）。"""

    def fin(key: str) -> float | None:
        v = m.get(key)
        if isinstance(v, (int, float)) and math.isfinite(float(v)):
            return float(v)
        return None

    return {
        "板块": sector,
        "代码": code,
        "简称": name,
        "ret_1m": fin("ret_1m"),
        "ret_3m": fin("ret_3m"),
        "ret_6m": fin("ret_6m"),
        "ret_1y": fin("ret_1y"),
        "ret_ytd": fin("ret_ytd"),
        "年化波动": fin("日收益年化波动率"),
        "最大回撤": fin("历史最大回撤"),
    }
